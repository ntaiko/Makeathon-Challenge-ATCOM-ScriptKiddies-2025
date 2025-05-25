import os
from getpass import getpass
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
from flask import Flask, render_template, request, jsonify
import logging
from typing import List, Optional, Union, Dict, Any

import openpyxl

import chromadb
from chromadb.utils import embedding_functions

from langchain_openai import ChatOpenAI, OpenAIEmbeddings
from langchain_core.prompts import PromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from pydantic import BaseModel, Field

load_dotenv()
if "OPENAI_API_KEY" not in os.environ:
    os.environ["OPENAI_API_KEY"] = getpass("Enter your OpenAI API Key: ")

UPLOAD_FOLDER = "uploads"
CHROMA_PERSIST_DIR = "chroma_prod_data_storage"
CHROMA_COLLECTION_NAME = "product_catalog_from_orders"
DEFAULT_DATA_FILE_PATH = 'static/file/data.xlsx'
EXCEL_PROCESSING_BATCH_SIZE = 2000
LLM_CONTEXT_PRODUCT_LIMIT = 30 

DEFAULT_ORDERS_SHEET = "orders"
ORDERS_ORDER_NUMBER_COLUMN = "OrderNumber"
ORDERS_SKU_COLUMN = "SKU"
ORDERS_ITEM_TITLE_COLUMN = "Item title"
ORDERS_CATEGORY_COLUMN_FOR_CONTEXT = "Category"
ORDERS_PRICE_COLUMN_FOR_CONTEXT = "FinalUnitPrice"
ORDERS_QUANTITY_COLUMN = "Quantity"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CHROMA_PERSIST_DIR, exist_ok=True)
os.makedirs(os.path.dirname(DEFAULT_DATA_FILE_PATH), exist_ok=True)

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

logging.basicConfig(level=logging.INFO)
logging.getLogger('werkzeug').setLevel(logging.WARNING)
logging.getLogger('langchain_community').setLevel(logging.WARNING)
logging.getLogger('langchain_openai').setLevel(logging.WARNING)
logging.getLogger('chromadb').setLevel(logging.INFO)
logging.getLogger('httpx').setLevel(logging.WARNING)
logging.getLogger('openai').setLevel(logging.WARNING) 
logging.getLogger('openapi_python_client').setLevel(logging.WARNING)


class BundleOutput(BaseModel):
    bundle_name: str = Field(..., description="Descriptive name for the product bundle (e.g., 'Summer Skincare Essentials Pack')")
    products: List[str] = Field(..., description="List of product names included in the bundle. These MUST be the full, exact product titles from the input context.")
    skus: List[str] = Field(..., description="List of Stock Keeping Units (SKUs) for each product in the 'products' list, in the same order. These should be the processed 'base' SKUs.")
    price_per_product: List[float] = Field(..., description="List of individual prices in euros (€) for each product in the 'products' list, *before* any bundle discount.")
    product_stock_levels: List[str] = Field(..., description="Notes on current stock levels for each product, in order.")
    product_sales_metrics: List[str] = Field(..., description="Key sales metrics or notes for each product, in order.")
    total_price: float = Field(..., description="Final discounted price of the entire bundle in euros (€).")
    original_total_price: float = Field(..., description="Total price of all products in the bundle if bought individually at their original prices, in euros (€).")
    discount_percent: float = Field(..., description="Percentage discount applied to the bundle.")
    trend: str = Field(..., description="The observed or targeted trend of the bundle.")
    margin: Union[float, str] = Field(..., description="The profit margin of the bundle. Can be a float (numeric value) or a string (e.g., 'Estimated due to lack of cost data') if costs are unavailable.")
    margin_type: Optional[str] = Field(default="percentage", description="Indicates if the margin is 'percentage' or 'absolute_eur'. Only relevant if margin is numeric.")
    summary: str = Field(..., description="A concise summary justifying the bundle, referencing data insights.")
    result: str = Field(..., description="A statement on the expected outcome or impact of this bundle.")
    recommended_duration_notes: Optional[str] = Field(default=None, description="Notes on recommended availability period based on stock or seasonality. Could include structured dates like 'Start:YYYY-MM-DD, End:YYYY-MM-DD'.")

lc_openai_embeddings = None
chroma_openai_ef = None
chroma_client = None
product_collection = None
model = None
chain = None

try:

    lc_openai_embeddings = OpenAIEmbeddings()
    app.logger.info("LangChain OpenAIEmbeddings initialized.")

    if "OPENAI_API_KEY" not in os.environ or not os.environ["OPENAI_API_KEY"]:
        raise ValueError("OPENAI_API_KEY not found or empty in environment variables for ChromaDB embedding function.")
    chroma_openai_ef = embedding_functions.OpenAIEmbeddingFunction(
        api_key=os.environ["OPENAI_API_KEY"],
        model_name="text-embedding-ada-002"
    )
    app.logger.info("ChromaDB OpenAIEmbeddingFunction initialized.")

    chroma_client = chromadb.PersistentClient(path=CHROMA_PERSIST_DIR)
    app.logger.info(f"ChromaDB PersistentClient initialized at {CHROMA_PERSIST_DIR}")
    product_collection = chroma_client.get_or_create_collection(
        name=CHROMA_COLLECTION_NAME,
        embedding_function=chroma_openai_ef
    )
    app.logger.info(f"ChromaDB collection '{CHROMA_COLLECTION_NAME}' loaded/created.")

    model = ChatOpenAI(temperature=0, model_name="gpt-4.1")
    prompt = PromptTemplate.from_template("""
    You are a data bot that creates bundle-ready JSON.
    Your goal is to help ecommerce managers create appealing bundles and price them effectively in euros (€) for a European context.

    Based on the provided product data context (which includes 'Product Name' for each item, associated with a 'BaseSKU'), generate a product bundle.
    The product data context is:
    {context}

    For the following user input, return a JSON object adhering to the specified schema.
    Crucially:
    1. The 'products' list in your JSON output MUST use the full, exact 'Product Name' values as found in the provided product data context for the BaseSKUs you select for the bundle. Do not abbreviate or alter these names.
    2. The 'products' list, 'skus' list (which should contain BaseSKUs), and 'price_per_product' list MUST have the same number of elements, and their order must correspond (e.g., price_per_product[i] is the price for products[i] which has skus[i] as its BaseSKU).
    3. If cost data for margin calculation is not available in the context, the 'margin' field in your JSON should be a string stating: 'Estimated due to lack of cost data'. Otherwise, provide a numeric margin.
    4. For 'recommended_duration_notes', if specific start and end dates can be inferred from stock levels, seasonality, or user request, try to provide them in a structured way like 'Start:YYYY-MM-DD, End:YYYY-MM-DD' or 'Duration: X days/weeks from YYYY-MM-DD'. Otherwise, provide general notes.
    5. Ensure all monetary values are in euros.

    User input: {user_input}

    JSON Schema to strictly follow:
    {{
    "bundle_name": "string (Descriptive name for the product bundle)",
    "products": "List[string] (List of the full, exact product names as found in the provided context for the chosen BaseSKUs)",
    "skus": "List[string] (List of BaseSKUs for each product, in the same order as 'products')",
    "price_per_product": "List[float] (List of individual prices in euros (€) for each product, before discount, in order)",
    "product_stock_levels": "List[string] (Notes on stock levels for each product, in order)",
    "product_sales_metrics": "List[string] (Key sales metrics for each product, in order)",
    "total_price": "float (Final discounted price of the bundle in euros (€))",
    "original_total_price": "float (Total original price of items if bought separately in euros (€))",
    "discount_percent": "float (Percentage discount for the bundle)",
    "trend": "string (The trend of the bundle)",
    "margin": "float OR string (The profit margin of the bundle. If numeric, it's a float. If estimated, it's a string like 'Estimated due to lack of cost data')",
    "margin_type": "string (Indicates if margin is 'percentage' or 'absolute_eur', default 'percentage'. Only relevant if margin is numeric.)",
    "summary": "string (Concise summary justifying the bundle, referencing data insights)",
    "result": "string (Expected outcome or impact of this bundle)",
    "recommended_duration_notes": "string (Optional: Notes on recommended availability period, considering stock levels. E.g., 'Offer for 2 weeks due to limited stock of Product Y', or 'Start: 2024-07-01, End: 2024-07-15')"
    }}
    """)
    parser = JsonOutputParser(pydantic_object=BundleOutput)
    chain = prompt | model | parser
    app.logger.info("Langchain LLM and chain initialized.")

except Exception as e:
    app.logger.critical(f"CRITICAL FAILURE during core component initialization: {e}", exc_info=True)

def get_column_indices_from_headers(headers: List[str], required_column_map: Dict[str, str]) -> Dict[str, Optional[int]]:
    indices = {}
    excel_header_to_internal_key = {v: k for k, v in required_column_map.items()}
    for i, header_val in enumerate(headers):
        header_key = str(header_val).strip() if header_val is not None else ""
        if header_key in excel_header_to_internal_key:
            indices[excel_header_to_internal_key[header_key]] = i

    for internal_key, excel_col_name in required_column_map.items():
        if internal_key not in indices:
            indices[internal_key] = None
            app.logger.warning(f"Required column for '{internal_key}' (expected Excel header: '{excel_col_name}') not found.")
    return indices

def process_and_ingest_excel_to_chroma(excel_file_path: str):
    if not product_collection:
        app.logger.error("ChromaDB product_collection is not available. Skipping ingestion.")
        return
    app.logger.info(f"Starting ingestion from Excel file: {excel_file_path}")
    try:
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        if DEFAULT_ORDERS_SHEET not in workbook.sheetnames:
            app.logger.error(f"Sheet '{DEFAULT_ORDERS_SHEET}' not found in '{excel_file_path}'. No data ingested.")
            return
        sheet = workbook[DEFAULT_ORDERS_SHEET]
    except FileNotFoundError:
        app.logger.error(f"Excel file not found at '{excel_file_path}'. No data ingested.")
        return
    except Exception as e:
        app.logger.error(f"Error opening or reading Excel file '{excel_file_path}': {e}", exc_info=True)
        return

    header_cells = sheet[1]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]

    column_map = {
        "order_number": ORDERS_ORDER_NUMBER_COLUMN, "complex_sku": ORDERS_SKU_COLUMN,
        "item_title": ORDERS_ITEM_TITLE_COLUMN, "category": ORDERS_CATEGORY_COLUMN_FOR_CONTEXT,
        "price": ORDERS_PRICE_COLUMN_FOR_CONTEXT, "quantity": ORDERS_QUANTITY_COLUMN,
    }
    col_indices = get_column_indices_from_headers(headers, column_map)

    if any(col_indices.get(key) is None for key in ["complex_sku", "item_title", "price"]):
        app.logger.error("Essential columns (SKU, Item title, Price) are missing from orders sheet. Aborting ingestion.")
        return

    processed_data_by_base_sku: Dict[str, Dict[str, Any]] = {}
    for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row_cells]
        def get_val(internal_key: str):
            idx = col_indices.get(internal_key)
            return row_values[idx] if idx is not None and idx < len(row_values) else None

        complex_sku_raw = get_val("complex_sku")
        complex_sku = str(complex_sku_raw).strip() if complex_sku_raw is not None else ""
        if not complex_sku: continue
        base_sku = complex_sku.split('|', 1)[0].strip()
        if not base_sku: continue

        item_title_raw = get_val("item_title")
        item_title = str(item_title_raw).strip() if item_title_raw and str(item_title_raw).strip() else f"Product {base_sku}"
        price_raw = get_val("price")
        try: price = float(price_raw) if price_raw is not None else 0.0
        except (ValueError, TypeError): price = 0.0; app.logger.debug(f"R{row_idx} (BSKU:{base_sku}): Price err '{price_raw}'->0.0")
        category_raw = get_val("category")
        category = str(category_raw).strip() if category_raw and str(category_raw).strip() else "N/A"
        quantity_raw = get_val("quantity")
        try: quantity = int(float(str(quantity_raw))) if quantity_raw is not None else 0
        except (ValueError, TypeError): quantity = 0; app.logger.debug(f"R{row_idx} (BSKU:{base_sku}): Qty err '{quantity_raw}'->0")
        order_number_raw = get_val("order_number")
        order_number = str(order_number_raw).strip() if order_number_raw else None

        if base_sku not in processed_data_by_base_sku:
            processed_data_by_base_sku[base_sku] = {
                "name": item_title, "price": price, "category": category,
                "total_quantity_sold": 0, "order_numbers": set(), "complex_sku_examples": set()
            }
        processed_data_by_base_sku[base_sku]["total_quantity_sold"] += quantity
        if order_number: processed_data_by_base_sku[base_sku]["order_numbers"].add(order_number)
        if complex_sku: processed_data_by_base_sku[base_sku]["complex_sku_examples"].add(complex_sku)

    docs_to_add, metadatas_to_add, ids_to_add = [], [], []
    for base_sku, data in processed_data_by_base_sku.items():
        num_orders = len(data["order_numbers"])
        total_sold = data["total_quantity_sold"]
        sales_note = f"Sales: {total_sold} units in {num_orders} orders."
        if total_sold > 100: sales_note += " (Popular)"
        elif 0 < total_sold < 10: sales_note += " (Slow Mover)"
        elif total_sold == 0: sales_note += " (No sales in this data)"
        doc_content = (f"Product: {data['name']}; BSKU: {base_sku}; Price:€{data['price']:.2f}; Cat:{data['category']}; {sales_note}")
        docs_to_add.append(doc_content)
        metadatas_to_add.append({
            "BaseSKU": base_sku, "ProductName": data["name"], "Price": data["price"],
            "Category": data["category"], "SalesMetrics": sales_note,
            "StockInfo": "Stock data N/A (placeholder)"
        })
        ids_to_add.append(base_sku)

    if docs_to_add:
        try:
            current_item_count = product_collection.count()
            if current_item_count > 0:
                app.logger.info(f"Attempting to clear {current_item_count} items from '{CHROMA_COLLECTION_NAME}'.")
                all_ids_in_collection = product_collection.get(limit=current_item_count, include=[])['ids']
                if all_ids_in_collection:
                    product_collection.delete(ids=all_ids_in_collection)
                    app.logger.info(f"Cleared {len(all_ids_in_collection)} items from collection.")
            
            app.logger.info(f"Prepared {len(docs_to_add)} total products for ingestion in batches.")
            
            num_batches = (len(docs_to_add) + EXCEL_PROCESSING_BATCH_SIZE - 1) // EXCEL_PROCESSING_BATCH_SIZE
            for i in range(num_batches):
                start_idx = i * EXCEL_PROCESSING_BATCH_SIZE
                end_idx = min((i + 1) * EXCEL_PROCESSING_BATCH_SIZE, len(docs_to_add))
                
                batch_docs = docs_to_add[start_idx:end_idx]
                batch_metadatas = metadatas_to_add[start_idx:end_idx]
                batch_ids = ids_to_add[start_idx:end_idx]
                
                app.logger.info(f"Ingesting batch {i+1}/{num_batches} ({len(batch_docs)} items) into ChromaDB.")
                product_collection.add(
                    documents=batch_docs,
                    metadatas=batch_metadatas,
                    ids=batch_ids
                )
                app.logger.info(f"Batch {i+1}/{num_batches} ingestion successful.")
            
            app.logger.info("All batches ingested into ChromaDB successfully.")
        except Exception as e_chroma:
            app.logger.error(f"Error during ChromaDB add/delete operation: {e_chroma}", exc_info=True)
    else:
        app.logger.info("No valid product data found in Excel to ingest into ChromaDB.")

def ensure_data_is_ingested(file_path=DEFAULT_DATA_FILE_PATH, force_reingest=False):
    if product_collection is None:
        app.logger.error("Cannot ensure data ingestion: ChromaDB collection not available.")
        return
    collection_is_empty = product_collection.count() == 0
    if force_reingest or collection_is_empty:
        action = "Re-ingesting" if force_reingest and not collection_is_empty else "Ingesting"
        app.logger.info(f"ChromaDB: {action} data from '{file_path}' (Force:{force_reingest}, Empty:{collection_is_empty}).")
        process_and_ingest_excel_to_chroma(file_path)
    else:
        app.logger.info(f"ChromaDB '{CHROMA_COLLECTION_NAME}' has {product_collection.count()} items. Default ingestion skipped.")

# --- Flask Routes ---
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/generate", methods=["POST"])
def generate_bundle_route():
    if not all([lc_openai_embeddings, product_collection, chain]):
        app.logger.error("Core component not available for generation (Embeddings, ChromaDB, or LLM Chain).")
        return jsonify({"error": "Server error: Core components not available. Check server logs."}), 500

    user_input = None
    if request.is_json:
        data = request.get_json(); user_input = data.get("user_input")
        app.logger.info("Received JSON request for bundle generation.")
    else: 
        user_input = request.form.get("user_input")
        file_obj = request.files.get("dataFile")
        app.logger.info("Received Form request for bundle generation.")
        if file_obj and file_obj.filename:
            filename = secure_filename(file_obj.filename)
        
            if filename.endswith((".xlsx", ".xls")):
                file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file_obj.save(file_path)
                app.logger.info(f"Uploaded Excel: {filename}. Data will be re-ingested into ChromaDB.")
                ensure_data_is_ingested(file_path, force_reingest=True) 
            else:
                app.logger.warning(f"Uploaded file '{filename}' is not an Excel file. Specific ingestion for this type is not yet implemented. Using existing ChromaDB data.")
        else:

            ensure_data_is_ingested(DEFAULT_DATA_FILE_PATH, force_reingest=False)

    if not user_input: return jsonify({"error": "User input (prompt) is required"}), 400

    try:
        num_items_in_collection = product_collection.count()
        if num_items_in_collection == 0:
            final_context_for_llm = ("No product data in ChromaDB. Upload Excel or check default load. "
                                     "Bundle based on general knowledge or user prompt details, or state data missing.")
            app.logger.warning("ChromaDB empty. LLM context minimal for bundle generation.")
        else:
            app.logger.info(f"Querying ChromaDB for products relevant to: '{user_input[:100]}...'")

            query_results = product_collection.query(
                query_texts=[user_input],
                n_results=min(LLM_CONTEXT_PRODUCT_LIMIT, num_items_in_collection), 
                include=["metadatas"]
            )

            if not query_results or not query_results['ids'] or not query_results['ids'][0]:
                 app.logger.warning("No relevant products found in ChromaDB for the query, or query failed. LLM context will be minimal.")
                 final_context_for_llm = "No specific products found matching your request. Please try a different prompt or ensure data is loaded."
            else:
                context_parts_from_chroma = ["Relevant product information from ChromaDB:"]
               
                retrieved_metadatas = query_results['metadatas'][0] if query_results['metadatas'] else []

                if not retrieved_metadatas:
                    app.logger.warning("Query returned results but no metadatas. LLM context will be minimal.")
                    final_context_for_llm = "Found some items but could not retrieve details. Please check data integrity."
                else:
                    for metadata_item in retrieved_metadatas:
                        base_sku = metadata_item.get("BaseSKU", "N/A")
                        product_name = metadata_item.get("ProductName", f"Product {base_sku}")
                        price_val = metadata_item.get("Price", 0.0)
                        category = metadata_item.get("Category", "N/A")
                        sales_metrics_note = metadata_item.get("SalesMetrics", "Sales data N/A")
                        stock_info_note = metadata_item.get("StockInfo", "Stock data N/A") 

                        price_display = f"€{price_val:.2f}" if isinstance(price_val, (int, float)) else "Price N/A"

                        context_parts_from_chroma.append(f"\n---\nProduct Name: {product_name} (BaseSKU: {base_sku})")
                        context_parts_from_chroma.append(f"Price: {price_display}")
                        context_parts_from_chroma.append(f"Category: {category}")
                        context_parts_from_chroma.append(f"Stock Info: {stock_info_note}")
                        context_parts_from_chroma.append(f"Sales Info: {sales_metrics_note}")
                    final_context_for_llm = "\n".join(context_parts_from_chroma)
        
        app.logger.info(f"Invoking LLM for bundle generation. User input: '{user_input[:100]}...'")
        app.logger.debug(f"Final context for LLM (first 400 chars): {final_context_for_llm[:400]}...")

        llm_result = chain.invoke({"user_input": user_input, "context": final_context_for_llm})
        app.logger.info("LLM invocation successful.")
        return jsonify(llm_result)
    except Exception as e:
        app.logger.error(f"Error during bundle generation: {e}", exc_info=True)
        return jsonify({"error": f"Internal error during bundle generation: {str(e)}"}), 500

if __name__ == "__main__":
    if not all([lc_openai_embeddings, chroma_client, product_collection, model, chain]):
        app.logger.critical("Core components failed init. App cannot run. Exiting.")
        exit(1)

    if not os.path.exists(DEFAULT_DATA_FILE_PATH):
        app.logger.info(f"Default data file {DEFAULT_DATA_FILE_PATH} not found. Creating dummy Excel.")
        try:
            headers = [
                ORDERS_ORDER_NUMBER_COLUMN, "CreatedDate", ORDERS_SKU_COLUMN, ORDERS_ITEM_TITLE_COLUMN,
                ORDERS_CATEGORY_COLUMN_FOR_CONTEXT, "Brand", ORDERS_QUANTITY_COLUMN,
                "OriginalUnitPrice", ORDERS_PRICE_COLUMN_FOR_CONTEXT, "OriginalLineTotal", "FinalLineTotal",
                "FinalOrderItemsTotal", "ShippingTotal", "TotalOrderAmount", "UserID"
            ]
            dummy_rows = [
                ["ORD01","2023-01-15","SKU001|RED|M","Eco Bottle (R,M)","Hydration","Eco",1,22.99,22.99,22.99,22.99,50.49,5,27.99,"UA"],
                ["ORD01","2023-01-15","SKU002|BLU|L","Org T-Shirt (B,L)","Apparel","AppCo","1.0",27.5,27.5,27.5,27.5,50.49,5,32.5,"UA"],
                ["ORD02","2023-01-16","SKU001|BLU|S","Eco Bottle (B,S)","Hydration","Eco","5.00",22.99,21,114.95,105,105,5,110,"UB"],
                ["ORD03","2023-01-17","SKU003|BLK|OS","Wireless Headphones Pro","Electronics","SoundMax","2",179.99,170.00,359.98,340.00,340.00,10.00,350.00,"UserC"],
                ["ORD04","2023-01-18","SKU_NO_TITLE|V1",None,"Gadgets","Generic","1.00",10.00,9.50,10.00,9.50,9.50,2.00,11.50,"UserD"],
                ["ORD05","2023-01-19","SKU001|GREEN|M","Eco Bottle (Green, M)","Hydration","Eco","3.00",22.99,22.99,68.97,68.97,68.97,5.00,73.97,"UserE"],
            ]
            wb = openpyxl.Workbook(); orders_ws = wb.active; orders_ws.title = DEFAULT_ORDERS_SHEET
            orders_ws.append(headers)
            for row in dummy_rows: orders_ws.append(row)
            wb.save(DEFAULT_DATA_FILE_PATH)
            app.logger.info(f"Created dummy Excel: {DEFAULT_DATA_FILE_PATH} with '{DEFAULT_ORDERS_SHEET}'.")
        except ImportError: app.logger.error("openpyxl not installed. Cannot create dummy .xlsx.")
        except Exception as e: app.logger.error(f"Could not create dummy Excel: {e}", exc_info=True)

    ensure_data_is_ingested(DEFAULT_DATA_FILE_PATH, force_reingest=True)
    app.run(debug=True, use_reloader=False)